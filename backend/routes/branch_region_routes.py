from flask import Blueprint, request, jsonify
from models.schema import db, Branch, Region, Executive, BranchExecutiveMap, RegionBranchMap

branch_region_bp = Blueprint("branch_region_bp", __name__)

### ---------------------- BRANCH ROUTES ----------------------

@branch_region_bp.route("/branches", methods=["GET"])
def get_branches():
    branches = Branch.query.order_by(Branch.name).all()
    return jsonify([{"id": b.id, "name": b.name} for b in branches])

@branch_region_bp.route("/branch", methods=["POST"])
def create_branch():
    name = request.json.get("name")
    if not name:
        return jsonify({"error": "Name required"}), 400
    if Branch.query.filter_by(name=name).first():
        return jsonify({"error": "Branch already exists"}), 400
    db.session.add(Branch(name=name))
    db.session.commit()
    return jsonify({"message": "Branch created"}), 201

@branch_region_bp.route("/branch/<string:name>", methods=["DELETE"])
def delete_branch(name):
    branch = Branch.query.filter_by(name=name).first()
    if not branch:
        return jsonify({"error": "Branch not found"}), 404
    BranchExecutiveMap.query.filter_by(branch_id=branch.id).delete()
    RegionBranchMap.query.filter_by(branch_id=branch.id).delete()
    db.session.delete(branch)
    db.session.commit()
    return jsonify({"message": "Branch deleted"}), 200

@branch_region_bp.route("/branch/<string:branch_name>/executives")
def get_execs_for_branch(branch_name):
    branch = Branch.query.filter_by(name=branch_name).first()
    if not branch:
        return jsonify([])
    maps = BranchExecutiveMap.query.filter_by(branch_id=branch.id).all()
    exec_ids = [m.executive_id for m in maps]
    execs = Executive.query.filter(Executive.id.in_(exec_ids)).all()
    return jsonify([e.name for e in execs])

### ---------------------- REGION ROUTES ----------------------

@branch_region_bp.route("/regions", methods=["GET"])
def get_regions():
    regions = Region.query.order_by(Region.name).all()
    return jsonify([{"id": r.id, "name": r.name} for r in regions])

@branch_region_bp.route("/region", methods=["POST"])
def create_region():
    name = request.json.get("name")
    if not name:
        return jsonify({"error": "Name required"}), 400
    if Region.query.filter_by(name=name).first():
        return jsonify({"error": "Region already exists"}), 400
    db.session.add(Region(name=name))
    db.session.commit()
    return jsonify({"message": "Region created"}), 201

@branch_region_bp.route("/region/<string:name>", methods=["DELETE"])
def delete_region(name):
    region = Region.query.filter_by(name=name).first()
    if not region:
        return jsonify({"error": "Region not found"}), 404
    RegionBranchMap.query.filter_by(region_id=region.id).delete()
    db.session.delete(region)
    db.session.commit()
    return jsonify({"message": "Region deleted"}), 200

@branch_region_bp.route("/region/<string:region_name>/branches")
def get_branches_for_region(region_name):
    region = Region.query.filter_by(name=region_name).first()
    if not region:
        return jsonify([])
    maps = RegionBranchMap.query.filter_by(region_id=region.id).all()
    branch_ids = [m.branch_id for m in maps]
    branches = Branch.query.filter(Branch.id.in_(branch_ids)).all()
    return jsonify([b.name for b in branches])

### ---------------------- MAPPING ROUTES ----------------------

@branch_region_bp.route("/map-branch-executives", methods=["POST"])
def map_branch_executives():
    data = request.json
    branch_name = data.get("branch")
    exec_names = data.get("executives", [])

    branch = Branch.query.filter_by(name=branch_name).first()
    if not branch:
        return jsonify({"error": "Branch not found"}), 404

    BranchExecutiveMap.query.filter_by(branch_id=branch.id).delete()

    for name in exec_names:
        exec_obj = Executive.query.filter_by(name=name).first()
        if exec_obj:
            db.session.add(BranchExecutiveMap(branch_id=branch.id, executive_id=exec_obj.id))

    db.session.commit()
    return jsonify({"message": "Mapping updated"}), 200

@branch_region_bp.route("/map-region-branches", methods=["POST"])
def map_region_branches():
    data = request.json
    region_name = data.get("region")
    branch_names = data.get("branches", [])

    region = Region.query.filter_by(name=region_name).first()
    if not region:
        return jsonify({"error": "Region not found"}), 404

    RegionBranchMap.query.filter_by(region_id=region.id).delete()

    for bname in branch_names:
        branch = Branch.query.filter_by(name=bname).first()
        if branch:
            db.session.add(RegionBranchMap(region_id=region.id, branch_id=branch.id))

    db.session.commit()
    return jsonify({"message": "Mapping updated"}), 200

### ---------------------- CURRENT MAPPING SUMMARY ----------------------

@branch_region_bp.route("/mappings", methods=["GET"])
def get_branch_region_exec_mapping():
    branches = Branch.query.all()
    result = []

    for b in branches:
        region = Region.query.join(RegionBranchMap, Region.id == RegionBranchMap.region_id)\
            .join(Branch, Branch.id == RegionBranchMap.branch_id)\
            .filter(Branch.id == b.id).first()

        execs = Executive.query.join(BranchExecutiveMap, Executive.id == BranchExecutiveMap.executive_id)\
            .filter(BranchExecutiveMap.branch_id == b.id).all()

        result.append({
            "branch": b.name,
            "region": region.name if region else "Unmapped",
            "executives": [e.name for e in execs],
            "count": len(execs)
        })

    return jsonify(result)

### ---------------------- UPLOAD BRANCH REGION FILE ----------------------

@branch_region_bp.route("/upload-branch-region-file", methods=["POST"])
def upload_branch_region_file():
    from io import BytesIO
    import pandas as pd

    file = request.files.get("file")
    sheet_name = request.form.get("sheet_name")
    header_row = int(request.form.get("header_row", 0))

    exec_code_col = request.form.get("exec_code_col") or None
    exec_name_col = request.form.get("exec_name_col")
    branch_col = request.form.get("branch_col")
    region_col = request.form.get("region_col") or None

    if not file or not sheet_name or not exec_name_col or not branch_col:
        return jsonify({"error": "Missing fields"}), 400

    df = pd.read_excel(BytesIO(file.read()), sheet_name=sheet_name, header=header_row)

    for _, row in df.iterrows():
        exec_name = str(row.get(exec_name_col)).strip()
        exec_code = str(row.get(exec_code_col)).strip() if exec_code_col else ""
        branch = str(row.get(branch_col)).strip()
        region = str(row.get(region_col)).strip() if region_col else ""

        if not exec_name or not branch:
            continue

        # Create or get executive
        exec_obj = Executive.query.filter_by(name=exec_name).first()
        if not exec_obj:
            exec_obj = Executive(name=exec_name, code=exec_code)
            db.session.add(exec_obj)
            db.session.flush()

        # Create or get branch
        branch_obj = Branch.query.filter_by(name=branch).first()
        if not branch_obj:
            branch_obj = Branch(name=branch)
            db.session.add(branch_obj)
            db.session.flush()

        # Create or get region
        region_obj = None
        if region:
            region_obj = Region.query.filter_by(name=region).first()
            if not region_obj:
                region_obj = Region(name=region)
                db.session.add(region_obj)
                db.session.flush()

        # Map executive to branch
        if not BranchExecutiveMap.query.filter_by(branch_id=branch_obj.id, executive_id=exec_obj.id).first():
            db.session.add(BranchExecutiveMap(branch_id=branch_obj.id, executive_id=exec_obj.id))

        # Map branch to region
        if region_obj and not RegionBranchMap.query.filter_by(branch_id=branch_obj.id, region_id=region_obj.id).first():
            db.session.add(RegionBranchMap(branch_id=branch_obj.id, region_id=region_obj.id))

    db.session.commit()
    return jsonify({"message": "File processed successfully."})


# Get sheet names from Excel file
@branch_region_bp.route("/get-sheet-names", methods=["POST"])
def get_sheet_names():
    from openpyxl import load_workbook
    file = request.files.get("file")
    wb = load_workbook(file)
    return jsonify({"sheets": wb.sheetnames})

# Preview column names from a sheet
@branch_region_bp.route("/preview-excel", methods=["POST"])
def preview_excel():
    from io import BytesIO
    import pandas as pd
    file = request.files.get("file")
    sheet_name = request.form.get("sheet_name")
    header_row = int(request.form.get("header_row", 0))
    df = pd.read_excel(BytesIO(file.read()), sheet_name=sheet_name, header=header_row)
    return jsonify({"columns": list(df.columns)})
