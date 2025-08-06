# routes/combined_data.py - Complete Updated Backend with Two Tables Per Sheet
from flask import Blueprint, request, jsonify, send_file, current_app, session
import pandas as pd
import numpy as np
from io import BytesIO
import traceback
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side,GradientFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, DataBarRule
import logging
import os
import re
import json
import hashlib
import uuid
import base64



# Create Blueprint
combined_bp = Blueprint('combined_data', __name__, url_prefix='/api/combined')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)



def copy_cell_comprehensive_formatting(source_cell, target_cell, excel_formatting=None):
    """
    Copy comprehensive formatting from source cell to target cell
    """
    try:
        # Copy basic value
        target_cell.value = source_cell.value
        
        # Font formatting with all properties
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name or 'Calibri',
                size=source_cell.font.size or 11,
                bold=source_cell.font.bold or False,
                italic=source_cell.font.italic or False,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline or 'none',
                strike=source_cell.font.strike or False,
                color=source_cell.font.color,
                charset=source_cell.font.charset,
                family=source_cell.font.family,
                scheme=source_cell.font.scheme
            )
        
        # Fill/Background formatting with gradient support
        if source_cell.fill and source_cell.fill.fill_type:
            if source_cell.fill.fill_type == 'solid':
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            elif source_cell.fill.fill_type in ['linear', 'path']:
                # Handle gradient fills
                try:
                    target_cell.fill = GradientFill(
                        type=source_cell.fill.type,
                        degree=getattr(source_cell.fill, 'degree', 0),
                        left=getattr(source_cell.fill, 'left', 0),
                        right=getattr(source_cell.fill, 'right', 0),
                        top=getattr(source_cell.fill, 'top', 0),
                        bottom=getattr(source_cell.fill, 'bottom', 0),
                        stop=getattr(source_cell.fill, 'stop', [])
                    )
                except:
                    # Fallback to pattern fill if gradient fails
                    target_cell.fill = PatternFill(
                        fill_type='solid',
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )
        
        # Enhanced alignment formatting
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation or 0,
                wrap_text=source_cell.alignment.wrap_text or False,
                shrink_to_fit=source_cell.alignment.shrink_to_fit or False,
                indent=source_cell.alignment.indent or 0,
                relativeIndent=getattr(source_cell.alignment, 'relativeIndent', 0),
                justifyLastLine=getattr(source_cell.alignment, 'justifyLastLine', False),
                readingOrder=getattr(source_cell.alignment, 'readingOrder', 0)
            )
        
        # Comprehensive border formatting
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom,
                diagonal=source_cell.border.diagonal,
                diagonal_direction=source_cell.border.diagonal_direction,
                outline=getattr(source_cell.border, 'outline', True),
                vertical=getattr(source_cell.border, 'vertical', None),
                horizontal=getattr(source_cell.border, 'horizontal', None)
            )
        
        # Enhanced number formatting with custom format preservation
        if source_cell.number_format and source_cell.number_format != 'General':
            target_cell.number_format = source_cell.number_format
        elif isinstance(source_cell.value, (int, float)) and not pd.isna(source_cell.value):
            # Apply formatting based on excel_formatting configuration
            if excel_formatting and excel_formatting.get('number_format', {}).get('force_format', False):
                format_pattern = excel_formatting.get('number_format', {}).get('format_pattern', '0.00')
                target_cell.number_format = format_pattern
            elif source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            else:
                # Default number formatting
                target_cell.number_format = '0.00'
        
        # Copy hyperlinks
        if hasattr(source_cell, 'hyperlink') and source_cell.hyperlink:
            target_cell.hyperlink = source_cell.hyperlink
        
        # Copy comments with formatting
        if hasattr(source_cell, 'comment') and source_cell.comment:
            target_cell.comment = source_cell.comment
        
        return True
        
    except Exception as e:
        logger.warning(f"Error copying comprehensive formatting for cell {source_cell.coordinate}: {str(e)}")
        # Fallback to basic value copy
        target_cell.value = source_cell.value
        return False

def copy_sheet_structural_formatting(source_sheet, target_sheet):
    """
    Copy structural formatting elements from source sheet to target sheet
    """
    try:
        # Copy column dimensions with enhanced precision
        for col_letter, col_dimension in source_sheet.column_dimensions.items():
            target_col_dim = target_sheet.column_dimensions[col_letter]
            
            # Copy width
            if col_dimension.width:
                target_col_dim.width = col_dimension.width
            
            # Copy visibility
            if hasattr(col_dimension, 'hidden') and col_dimension.hidden:
                target_col_dim.hidden = col_dimension.hidden
            
            # Copy auto size
            if hasattr(col_dimension, 'auto_size') and col_dimension.auto_size:
                target_col_dim.auto_size = col_dimension.auto_size
            
            # Copy outline level
            if hasattr(col_dimension, 'outline_level') and col_dimension.outline_level:
                target_col_dim.outline_level = col_dimension.outline_level
        
        # Copy row dimensions with enhanced properties
        for row_num, row_dimension in source_sheet.row_dimensions.items():
            target_row_dim = target_sheet.row_dimensions[row_num]
            
            # Copy height
            if row_dimension.height:
                target_row_dim.height = row_dimension.height
            
            # Copy visibility
            if hasattr(row_dimension, 'hidden') and row_dimension.hidden:
                target_row_dim.hidden = row_dimension.hidden
            
            # Copy outline level
            if hasattr(row_dimension, 'outline_level') and row_dimension.outline_level:
                target_row_dim.outline_level = row_dimension.outline_level
        
        # Copy merged cells ranges
        for merged_range in source_sheet.merged_cells.ranges:
            try:
                target_sheet.merge_cells(str(merged_range))
            except Exception as merge_error:
                logger.warning(f"Could not merge cells {merged_range}: {str(merge_error)}")
        
        # Copy freeze panes
        if source_sheet.freeze_panes:
            target_sheet.freeze_panes = source_sheet.freeze_panes
        
        # Copy print settings
        try:
            target_sheet.page_setup.orientation = source_sheet.page_setup.orientation
            target_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
            target_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
            target_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
            target_sheet.page_setup.scale = source_sheet.page_setup.scale
            
            # Copy margins
            if source_sheet.page_margins:
                target_sheet.page_margins.left = source_sheet.page_margins.left
                target_sheet.page_margins.right = source_sheet.page_margins.right
                target_sheet.page_margins.top = source_sheet.page_margins.top
                target_sheet.page_margins.bottom = source_sheet.page_margins.bottom
                target_sheet.page_margins.header = source_sheet.page_margins.header
                target_sheet.page_margins.footer = source_sheet.page_margins.footer
                
        except Exception as print_error:
            logger.warning(f"Could not copy print settings: {str(print_error)}")
        
        # Copy conditional formatting
        try:
            for cf_range, cf_rules in source_sheet.conditional_formatting._cf_rules.items():
                for rule in cf_rules:
                    target_sheet.conditional_formatting.add(cf_range, rule)
        except Exception as cf_error:
            logger.warning(f"Could not copy conditional formatting: {str(cf_error)}")
        
        # Copy data validations
        try:
            for dv in source_sheet.data_validations.dataValidation:
                target_sheet.add_data_validation(dv)
        except Exception as dv_error:
            logger.warning(f"Could not copy data validations: {str(dv_error)}")
            
        return True
        
    except Exception as e:
        logger.error(f"Error copying structural formatting: {str(e)}")
        return False

def apply_intelligent_number_formatting(worksheet, excel_formatting=None):
    """
    Apply intelligent number formatting to a worksheet based on content
    """
    if not excel_formatting:
        return
    
    number_format_config = excel_formatting.get('number_format', {})
    if not number_format_config.get('apply_to_all_numbers', False):
        return
    
    format_pattern = number_format_config.get('format_pattern', '0.00')
    force_format = number_format_config.get('force_format', False)
    
    try:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    # Check if cell already has a specific number format
                    if force_format or cell.number_format == 'General':
                        cell.number_format = format_pattern
                        
        logger.info(f"Applied intelligent number formatting '{format_pattern}' to worksheet {worksheet.title}")
        
    except Exception as e:
        logger.warning(f"Could not apply intelligent number formatting: {str(e)}")

def detect_and_preserve_callback_formatting(file_metadata):
    """
    Detect if a file was generated via callback and should preserve special formatting
    """
    if not file_metadata:
        return False
    
    # Check for callback indicators
    callback_indicators = [
        file_metadata.get('wasCallbackStored', False),
        file_metadata.get('autoStored', False),
        file_metadata.get('originalFormatting', False),
        'callback' in str(file_metadata.get('source', '')).lower(),
        file_metadata.get('formattingApplied', False)
    ]
    
    return any(callback_indicators)

def enhance_overview_sheet_formatting(overview_sheet):
    """
    Apply enhanced formatting to the overview sheet
    """
    try:
        # Define enhanced color scheme
        colors = {
            'primary': "1F4E79",      # Dark blue
            'secondary': "4472C4",     # Medium blue  
            'success': "28A745",       # Green
            'info': "17A2B8",          # Teal
            'warning': "FFC107",       # Yellow
            'light': "F8F9FA",         # Light gray
            'white': "FFFFFF"          # White
        }
        
        # Apply alternating row colors to data sections
        current_row = 1
        for row in overview_sheet.iter_rows():
            if current_row > 3:  # Skip header rows
                for cell in row:
                    if cell.value and current_row % 2 == 0:
                        # Apply light background to even rows
                        cell.fill = PatternFill(
                            start_color=colors['light'], 
                            end_color=colors['light'], 
                            fill_type="solid"
                        )
            current_row += 1
        
        # Add borders to all data cells
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for row in overview_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
        
        logger.info("Enhanced overview sheet formatting applied successfully")
        
    except Exception as e:
        logger.warning(f"Could not enhance overview sheet formatting: {str(e)}")

# Global callback registry for file storage
FILE_STORAGE_CALLBACKS = {}

def register_file_storage_callback(callback_id, callback_info):
    """Register a callback for automatic file storage"""
    FILE_STORAGE_CALLBACKS[callback_id] = {
        'callback_info': callback_info,
        'timestamp': datetime.now().isoformat(),
        'status': 'active'
    }
    logger.info(f"File storage callback registered: {callback_id}")

def trigger_file_storage_callback(callback_id, file_data):
    """Trigger a registered file storage callback"""
    if callback_id in FILE_STORAGE_CALLBACKS:
        callback = FILE_STORAGE_CALLBACKS[callback_id]
        callback['last_triggered'] = datetime.now().isoformat()
        callback['files_stored'] = callback.get('files_stored', 0) + 1
        logger.info(f"File storage callback triggered: {callback_id}, File: {file_data.get('name', 'Unknown')}")
        return True
    return False

def generate_file_metadata(file_data, analysis_type, source_info=None):
    """Generate comprehensive metadata for stored files"""
    file_size = len(file_data) if isinstance(file_data, bytes) else 0
    
    metadata = {
        'id': str(uuid.uuid4()),
        'name': f"{analysis_type}_combined_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        'type': f'{analysis_type}-combined-analysis',
        'size': file_size,
        'timestamp': datetime.now().isoformat(),
        'analysis_type': analysis_type,
        'auto_stored': True,
        'source_info': source_info or {},
        'hash': hashlib.md5(file_data).hexdigest() if isinstance(file_data, bytes) else None
    }
    
    return metadata

@combined_bp.route('/register-storage-callback', methods=['POST'])
def register_storage_callback():
    """Register a callback for automatic file storage"""
    try:
        data = request.get_json()
        
        callback_id = data.get('callback_id') or str(uuid.uuid4())
        analysis_type = data.get('analysis_type', 'combined')
        component_name = data.get('component_name', 'Unknown')
        
        callback_info = {
            'analysis_type': analysis_type,
            'component_name': component_name,
            'auto_store': data.get('auto_store', True),
            'file_naming_pattern': data.get('file_naming_pattern', f'{analysis_type}_{{timestamp}}'),
            'metadata_fields': data.get('metadata_fields', {}),
            'storage_options': data.get('storage_options', {})
        }
        
        register_file_storage_callback(callback_id, callback_info)
        
        return jsonify({
            'success': True,
            'callback_id': callback_id,
            'message': f'Storage callback registered for {analysis_type} analysis',
            'callback_info': callback_info
        })
        
    except Exception as e:
        logger.error(f"Error registering storage callback: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to register storage callback: {str(e)}'
        }), 500

@combined_bp.route('/validate-data', methods=['POST'])
def validate_data():
    """Validate data before processing - Enhanced for all analysis types"""
    try:
        data = request.get_json()
        
        mt_data = data.get('mt_data', [])
        value_data = data.get('value_data', [])
        fiscal_year = data.get('fiscal_year', '')
        stored_files_count = data.get('stored_files_count', 0)
        
        # Enhanced: All analysis types
        sales_analysis_files = data.get('sales_analysis_files', [])
        ero_pw_files = data.get('ero_pw_files', [])
        ts_pw_files = data.get('ts_pw_files', [])
        product_files = data.get('product_files', [])
        region_files = data.get('region_files', [])
        
        validation_results = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'info': {
                'mt_records': len(mt_data),
                'value_records': len(value_data),
                'fiscal_year': fiscal_year,
                'stored_files': stored_files_count,
                'sales_analysis_files': len(sales_analysis_files),
                'ero_pw_files': len(ero_pw_files),
                'ts_pw_files': len(ts_pw_files),
                'product_files': len(product_files),
                'region_files': len(region_files)
            }
        }
        
        # Validate each analysis type
        if sales_analysis_files:
            sales_analysis_mt_count = sum(1 for f in sales_analysis_files if 'mt' in f.get('type', '').lower())
            sales_analysis_value_count = sum(1 for f in sales_analysis_files if 'value' in f.get('type', '').lower())
            
            validation_results['info']['sales_analysis_breakdown'] = {
                'mt_files': sales_analysis_mt_count,
                'value_files': sales_analysis_value_count,
                'excel_files': len([f for f in sales_analysis_files if f.get('type') == 'sales-analysis-excel'])
            }
        
        # Validate MT data
        if not mt_data:
            validation_results['warnings'].append('No MT data provided')
        
        # Validate Value data
        if not value_data:
            validation_results['warnings'].append('No Value data provided')
        
        # Validate fiscal year
        if not fiscal_year:
            validation_results['warnings'].append('No fiscal year provided')
        
        # Check total data sources
        total_data_sources = (len(mt_data) + len(value_data) + len(ero_pw_files) + 
                            len(ts_pw_files) + len(product_files) + len(region_files) + 
                            len(sales_analysis_files))
        
        if total_data_sources == 0:
            validation_results['errors'].append('No data provided for validation')
            validation_results['valid'] = False
        
        logger.info(f"Enhanced data validation completed: {validation_results}")
        
        return jsonify({
            'success': True,
            'validation': validation_results
        })
        
    except Exception as e:
        logger.error(f"Error validating data: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Validation failed: {str(e)}'
        }), 500
# Enhanced backend modification for category-based Excel generation

@combined_bp.route('/generate-master-excel-with-callback', methods=['POST'])
def generate_master_excel_with_callback():
    """Generate master combined Excel file with one sheet per file category - NO OVERVIEW VERSION"""
    try:
        data = request.get_json()
        
        # Extract data from request
        mt_data = data.get('mt_data', [])
        value_data = data.get('value_data', [])
        fiscal_year = data.get('fiscal_year', 'Unknown')
        file_name = data.get('file_name', f'category_merged_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
        
        # Category-specific data
        sales_analysis_files = data.get('sales_analysis_files', [])
        sales_analysis_mt_data = data.get('sales_analysis_mt_data', [])
        sales_analysis_value_data = data.get('sales_analysis_value_data', [])
        
        product_analysis_files = data.get('product_analysis_files', [])
        
        ero_pw_files_info = data.get('ero_pw_files_info', [])
        ero_pw_mt_data = data.get('ero_pw_mt_data', [])
        ero_pw_value_data = data.get('ero_pw_value_data', [])
        
        ts_pw_files = data.get('ts_pw_files', [])
        
        combined_files_info = data.get('combined_files_info', [])
        other_files_info = data.get('other_files_info', [])
        
        # Callback configuration
        callback_id = data.get('callback_id')
        auto_store = data.get('auto_store', True)
        
        logger.info(f"Generating category-based Excel (no overview): {file_name}")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define styles for dual tables
        # MT Table styles (Green theme)
        mt_title_font = Font(bold=True, size=14, color="FFFFFF")
        mt_title_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        mt_header_fill = PatternFill(start_color="5CB85C", end_color="5CB85C", fill_type="solid")
        
        # Value Table styles (Blue theme)
        value_title_font = Font(bold=True, size=14, color="FFFFFF")
        value_title_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        value_header_fill = PatternFill(start_color="5BC0DE", end_color="5BC0DE", fill_type="solid")
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_alignment_left = Alignment(horizontal="left", vertical="center")
        data_alignment_right = Alignment(horizontal="right", vertical="center")
        data_alignment_center = Alignment(horizontal="center", vertical="center")
        
        total_font = Font(bold=True, color="155724")
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def create_dual_table_sheet(sheet_name, mt_df, value_df, sheet_index):
            """Create a sheet with exactly 2 tables - MT and Value"""
            ws = wb.create_sheet(sheet_name, sheet_index - 1)
            
            current_row = 1
            
            # MT DATA TABLE
            if not mt_df.empty:
                # MT Title
                mt_title = ws.cell(row=current_row, column=1, value="MT DATA")
                mt_title.font = mt_title_font
                mt_title.fill = mt_title_fill
                mt_title.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(mt_df.columns))
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # MT Headers
                for col_idx, column in enumerate(mt_df.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=column)
                    cell.font = header_font
                    cell.fill = mt_header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # MT Data rows
                for row_data in mt_df.itertuples(index=False):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        
                        if col_idx == 1:  # First column
                            cell.alignment = data_alignment_left
                            # Check for total rows
                            if isinstance(value, str) and ('TOTAL' in str(value).upper() or 'ACCLLP' in str(value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                        else:
                            if isinstance(value, (int, float)) and not pd.isna(value):
                                cell.number_format = '#,##0.00'
                                cell.alignment = data_alignment_right
                            else:
                                cell.alignment = data_alignment_center
                            
                            # Highlight total rows
                            first_col_value = row_data[0] if len(row_data) > 0 else ""
                            if isinstance(first_col_value, str) and ('TOTAL' in str(first_col_value).upper() or 'ACCLLP' in str(first_col_value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                    
                    current_row += 1
                
                current_row += 2  # Space between tables
            
            # VALUE DATA TABLE
            if not value_df.empty:
                # Value Title
                value_title = ws.cell(row=current_row, column=1, value="VALUE DATA")
                value_title.font = value_title_font
                value_title.fill = value_title_fill
                value_title.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(value_df.columns))
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # Value Headers
                for col_idx, column in enumerate(value_df.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=column)
                    cell.font = header_font
                    cell.fill = value_header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # Value Data rows
                for row_data in value_df.itertuples(index=False):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        
                        if col_idx == 1:  # First column
                            cell.alignment = data_alignment_left
                            # Check for total rows
                            if isinstance(value, str) and ('TOTAL' in str(value).upper() or 'ACCLLP' in str(value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                        else:
                            if isinstance(value, (int, float)) and not pd.isna(value):
                                cell.number_format = '#,##0.00'
                                cell.alignment = data_alignment_right
                            else:
                                cell.alignment = data_alignment_center
                            
                            # Highlight total rows
                            first_col_value = row_data[0] if len(row_data) > 0 else ""
                            if isinstance(first_col_value, str) and ('TOTAL' in str(first_col_value).upper() or 'ACCLLP' in str(first_col_value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                    
                    current_row += 1
            
            # If both tables are empty, add a message
            if mt_df.empty and value_df.empty:
                ws.cell(row=1, column=1, value="No data available for this category")
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            
            # Auto-adjust column widths
            max_cols = max(
                len(mt_df.columns) if not mt_df.empty else 5,
                len(value_df.columns) if not value_df.empty else 5
            )
            
            for col_idx in range(1, max_cols + 1):
                if col_idx == 1:  # First column - wider
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35
                else:  # Data columns
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            
            # Freeze panes at A3 (below titles and headers)
            ws.freeze_panes = 'A3'
            
            return ws
        
        # Create sheets in the specific order requested
        sheet_count = 0
        created_sheets = []
        
        # Always create these 5 sheets in this exact order, even if empty
        
        # 1. SalesMonthwise Sheet (1st sheet)
        sheet_count += 1
        sales_mt_df = pd.DataFrame(sales_analysis_mt_data) if sales_analysis_mt_data else pd.DataFrame()
        sales_value_df = pd.DataFrame(sales_analysis_value_data) if sales_analysis_value_data else pd.DataFrame()
        create_dual_table_sheet("SalesMonthwise", sales_mt_df, sales_value_df, sheet_count)
        created_sheets.append("SalesMonthwise")
        logger.info(f"Created SalesMonthwise sheet - MT: {len(sales_mt_df)} rows, Value: {len(sales_value_df)} rows")
        
        # 2. Region Sheet (2nd sheet)
        sheet_count += 1
        region_mt_df = pd.DataFrame(mt_data) if mt_data else pd.DataFrame()
        region_value_df = pd.DataFrame(value_data) if value_data else pd.DataFrame()
        create_dual_table_sheet("Region", region_mt_df, region_value_df, sheet_count)
        created_sheets.append("Region")
        logger.info(f"Created Region sheet - MT: {len(region_mt_df)} rows, Value: {len(region_value_df)} rows")
        
        # 3. Product Sheet (3rd sheet)
        sheet_count += 1
        # Use empty dataframes for now - you can populate with actual product data if available
        product_mt_df = pd.DataFrame()
        product_value_df = pd.DataFrame()
        create_dual_table_sheet("Product", product_mt_df, product_value_df, sheet_count)
        created_sheets.append("Product")
        logger.info("Created Product sheet")
        
        # 4. TS-PW Sheet (4th sheet)
        sheet_count += 1
        # Use empty dataframes for now - you can populate with actual TS-PW data if available
        ts_mt_df = pd.DataFrame()
        ts_value_df = pd.DataFrame()
        create_dual_table_sheet("TS-PW", ts_mt_df, ts_value_df, sheet_count)
        created_sheets.append("TS-PW")
        logger.info("Created TS-PW sheet")
        
        # 5. ERO-PW Sheet (5th sheet)
        sheet_count += 1
        ero_mt_df = pd.DataFrame(ero_pw_mt_data) if ero_pw_mt_data else pd.DataFrame()
        ero_value_df = pd.DataFrame(ero_pw_value_data) if ero_pw_value_data else pd.DataFrame()
        create_dual_table_sheet("ERO-PW", ero_mt_df, ero_value_df, sheet_count)
        created_sheets.append("ERO-PW")
        logger.info(f"Created ERO-PW sheet - MT: {len(ero_mt_df)} rows, Value: {len(ero_value_df)} rows")
        
        # Note: We always create all 5 sheets, so sheet_count will always be 5
        
        # Set print settings for all sheets
        for sheet in wb.worksheets:
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.getvalue()
        
        # Generate file metadata
        source_info = {
            'fiscal_year': fiscal_year,
            'sheets_created': len(created_sheets),
            'format_type': 'category_based_dual_tables',
            'created_sheets': created_sheets
        }
        
        file_metadata = generate_file_metadata(file_data, 'category_based_merged', source_info)
        file_metadata['name'] = f"{file_name}.xlsx"
        
        # Trigger callback for automatic storage if configured
        storage_triggered = False
        if auto_store and callback_id:
            storage_triggered = trigger_file_storage_callback(callback_id, file_metadata)
        
        logger.info(f"Successfully generated category-based merged Excel: {file_name}.xlsx - {len(created_sheets)} sheets")
        
        # Reset BytesIO for download
        output = BytesIO(file_data)
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{file_name}.xlsx"
        )
        
        # Add custom headers
        response.headers['X-Storage-Callback-Triggered'] = str(storage_triggered)
        response.headers['X-Sheets-Created'] = str(len(created_sheets))
        response.headers['X-Categories-Included'] = str(len(created_sheets))
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating category-based merged Excel: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Failed to generate category-based merged Excel: {str(e)}'
        }), 500

@combined_bp.route('/generate-selected-excel', methods=['POST'])
def generate_selected_excel():
    """Generate combined Excel from selected files with dual tables per sheet"""
    try:
        data = request.get_json()
        
        # Extract data from request
        mt_data = data.get('mt_data', [])
        value_data = data.get('value_data', [])
        fiscal_year = data.get('fiscal_year', 'Unknown')
        file_name = data.get('file_name', f'selected_combined_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
        
        # Selected files info
        selected_files_info = data.get('selected_files_info', [])
        
        # Sales Analysis selected
        sales_analysis_selected = data.get('sales_analysis_selected', False)
        sales_analysis_mt_data = data.get('sales_analysis_mt_data', [])
        sales_analysis_value_data = data.get('sales_analysis_value_data', [])
        
        # ERO-PW selected
        ero_pw_selected = data.get('ero_pw_selected', False)
        ero_pw_mt_data = data.get('ero_pw_mt_data', [])
        ero_pw_value_data = data.get('ero_pw_value_data', [])
        
        # Callback configuration
        callback_id = data.get('callback_id')
        auto_store = data.get('auto_store', True)
        
        logger.info(f"Generating selected combined Excel with dual tables: {file_name}")
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define styles (same as master Excel)
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # MT Table styles (Green)
        mt_title_font = Font(bold=True, size=14, color="FFFFFF")
        mt_title_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        
        # Value Table styles (Blue)
        value_title_font = Font(bold=True, size=14, color="FFFFFF")
        value_title_fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        total_font = Font(bold=True, color="155724")
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def format_selected_dual_table_worksheet(ws, mt_df, value_df, sheet_title, analysis_type=""):
            """Apply formatting to selected worksheet with two tables (MT and Value) stacked vertically"""
            
            current_row = 1
            
            # Add main sheet title
            if not mt_df.empty or not value_df.empty:
                max_cols = max(
                    len(mt_df.columns) if not mt_df.empty else 0, 
                    len(value_df.columns) if not value_df.empty else 0,
                    5  # Minimum columns for title
                )
                
                title_cell = ws.cell(row=current_row, column=1, value=sheet_title)
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = title_alignment
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
                ws.row_dimensions[current_row].height = 30
                current_row += 2
            
            # MT Table Section
            if not mt_df.empty:
                # MT Table Title
                mt_title_cell = ws.cell(row=current_row, column=1, value=f"{analysis_type} - MT DATA (SELECTED)")
                mt_title_cell.font = mt_title_font
                mt_title_cell.fill = mt_title_fill
                mt_title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(mt_df.columns))
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # MT Headers
                for col_idx, column in enumerate(mt_df.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=column)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # MT Data
                for row_data in mt_df.itertuples(index=False):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        
                        # Format based on column type
                        if col_idx == 1:  # First column
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            # Check for total rows
                            if isinstance(value, str) and ('TOTAL' in str(value).upper() or 'ACCLLP' in str(value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                        else:  # Data columns
                            if isinstance(value, (int, float)) and not pd.isna(value):
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            # Highlight total rows
                            first_col_value = row_data[0] if len(row_data) > 0 else ""
                            if isinstance(first_col_value, str) and ('TOTAL' in str(first_col_value).upper() or 'ACCLLP' in str(first_col_value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                    current_row += 1
                
                current_row += 2  # Space between tables
            
            # Value Table Section
            if not value_df.empty:
                # Value Table Title
                value_title_cell = ws.cell(row=current_row, column=1, value=f"{analysis_type} - VALUE DATA (SELECTED)")
                value_title_cell.font = value_title_font
                value_title_cell.fill = value_title_fill
                value_title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(value_df.columns))
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # Value Headers
                for col_idx, column in enumerate(value_df.columns, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=column)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                
                # Value Data
                for row_data in value_df.itertuples(index=False):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        
                        # Format based on column type
                        if col_idx == 1:  # First column
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            # Check for total rows
                            if isinstance(value, str) and ('TOTAL' in str(value).upper() or 'ACCLLP' in str(value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                        else:  # Data columns
                            if isinstance(value, (int, float)) and not pd.isna(value):
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            # Highlight total rows
                            first_col_value = row_data[0] if len(row_data) > 0 else ""
                            if isinstance(first_col_value, str) and ('TOTAL' in str(first_col_value).upper() or 'ACCLLP' in str(first_col_value).upper()):
                                cell.font = total_font
                                cell.fill = total_fill
                    current_row += 1
            
            # Auto-adjust column widths
            max_cols = max(
                len(mt_df.columns) if not mt_df.empty else 0, 
                len(value_df.columns) if not value_df.empty else 0
            )
            for col_idx in range(1, max_cols + 1):
                if col_idx == 1:  # First column - wider
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35
                else:  # Data columns
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14
            
            # Add freeze panes at the first data row
            if not mt_df.empty:
                ws.freeze_panes = ws['B6']
            elif not value_df.empty:
                ws.freeze_panes = ws['B4']
        
        sheet_count = 0
        
        # Create sheets based on selected data
        
        # Region Analysis (if region data available)
        if mt_data or value_data:
            try:
                sheet_count += 1
                region_mt_df = pd.DataFrame(mt_data) if mt_data else pd.DataFrame()
                region_value_df = pd.DataFrame(value_data) if value_data else pd.DataFrame()
                region_sheet = wb.create_sheet(f"{sheet_count:02d}_Region_Selected", sheet_count - 1)
                format_selected_dual_table_worksheet(region_sheet, region_mt_df, region_value_df, 
                                                   "REGION ANALYSIS - SELECTED", "REGION ANALYSIS")
                logger.info(f"Created Selected Region Analysis sheet with MT: {len(region_mt_df)}, Value: {len(region_value_df)} rows")
            except Exception as e:
                logger.error(f"Error creating Selected Region Analysis sheet: {str(e)}")
        
        # Sales Analysis (if selected)
        if sales_analysis_selected and (sales_analysis_mt_data or sales_analysis_value_data):
            try:
                sheet_count += 1
                sales_mt_df = pd.DataFrame(sales_analysis_mt_data) if sales_analysis_mt_data else pd.DataFrame()
                sales_value_df = pd.DataFrame(sales_analysis_value_data) if sales_analysis_value_data else pd.DataFrame()
                sales_sheet = wb.create_sheet(f"{sheet_count:02d}_Sales_Selected", sheet_count - 1)
                format_selected_dual_table_worksheet(sales_sheet, sales_mt_df, sales_value_df, 
                                                   "SALES ANALYSIS - SELECTED", "SALES ANALYSIS")
                logger.info(f"Created Selected Sales Analysis sheet with MT: {len(sales_mt_df)}, Value: {len(sales_value_df)} rows")
            except Exception as e:
                logger.error(f"Error creating Selected Sales Analysis sheet: {str(e)}")
        
        # ERO-PW Analysis (if selected)
        if ero_pw_selected and (ero_pw_mt_data or ero_pw_value_data):
            try:
                sheet_count += 1
                ero_pw_mt_df = pd.DataFrame(ero_pw_mt_data) if ero_pw_mt_data else pd.DataFrame()
                ero_pw_value_df = pd.DataFrame(ero_pw_value_data) if ero_pw_value_data else pd.DataFrame()
                ero_pw_sheet = wb.create_sheet(f"{sheet_count:02d}_ERO_PW_Selected", sheet_count - 1)
                format_selected_dual_table_worksheet(ero_pw_sheet, ero_pw_mt_df, ero_pw_value_df, 
                                                   "ERO-PW ANALYSIS - SELECTED", "ERO-PW ANALYSIS")
                logger.info(f"Created Selected ERO-PW Analysis sheet with MT: {len(ero_pw_mt_df)}, Value: {len(ero_pw_value_df)} rows")
            except Exception as e:
                logger.error(f"Error creating Selected ERO-PW Analysis sheet: {str(e)}")
        
        # If no sheets were created, create a default message sheet
        if sheet_count == 0:
            default_sheet = wb.create_sheet("No_Selected_Data", 0)
            default_sheet.cell(row=1, column=1, value="No selected data available for Excel generation")
            default_sheet.cell(row=2, column=1, value="Please select files and ensure data is available")
            default_sheet.cell(row=3, column=1, value=f"Selected files: {len(selected_files_info)}")
            logger.warning("No selected data available - created default message sheet")
        
        # Set print settings for all sheets
        for sheet in wb.worksheets:
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.getvalue()
        
        # Generate file metadata for storage
        source_info = {
            'selected_files': len(selected_files_info),
            'fiscal_year': fiscal_year,
            'total_records': len(mt_data) + len(value_data) + len(sales_analysis_mt_data) + len(sales_analysis_value_data) + len(ero_pw_mt_data) + len(ero_pw_value_data),
            'sheets_created': sheet_count,
            'format_type': 'selected_dual_tables_per_sheet',
            'integrations': {
                'sales_analysis': sales_analysis_selected,
                'ero_pw': ero_pw_selected
            }
        }
        
        file_metadata = generate_file_metadata(file_data, 'selected_dual_table_combined', source_info)
        file_metadata['name'] = f"{file_name}.xlsx"
        
        # Trigger callback for automatic storage if configured
        storage_triggered = False
        if auto_store and callback_id:
            storage_triggered = trigger_file_storage_callback(callback_id, file_metadata)
        
        logger.info(f"Successfully generated selected combined Excel file with dual tables: {file_name}.xlsx - {sheet_count} sheets (Storage callback: {'triggered' if storage_triggered else 'not configured'})")
        
        # Reset BytesIO for download
        output = BytesIO(file_data)
        
        response = send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{file_name}.xlsx"
        )
        
        # Add custom headers for callback information
        response.headers['X-File-Metadata'] = json.dumps(file_metadata)
        response.headers['X-Storage-Callback-Triggered'] = str(storage_triggered)
        response.headers['X-Selected-Files'] = str(len(selected_files_info))
        response.headers['X-Sheets-Created'] = str(sheet_count)
        response.headers['X-Format-Type'] = 'selected_dual_tables_per_sheet'
        
        return response
        
    except Exception as e:
        logger.error(f"Error generating selected combined Excel: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Failed to generate selected combined Excel: {str(e)}'
        }), 500
    
# Add this route to your combined_data.py file

@combined_bp.route('/combine-category-excel-files', methods=['POST'])
def combine_category_excel_files():
    """Combine all Excel files from a specific category into one master file with all sheets"""
    try:
        data = request.get_json()
        
        category_name = data.get('category_name', 'unknown')
        file_data_list = data.get('files', [])
        fiscal_year = data.get('fiscal_year', '')
        region_mt_columns = data.get('region_mt_columns', [])
        region_value_columns = data.get('region_value_columns', [])
        combination_type = data.get('combination_type', 'category_based')
        metadata = data.get('metadata', {})
        
        if not file_data_list:
            return jsonify({
                'success': False,
                'error': 'No files provided for combining'
            }), 400

        logger.info(f"Combining {len(file_data_list)} files from {category_name} category")

        # Create a new workbook
        combined_wb = openpyxl.Workbook()
        combined_wb.remove(combined_wb.active)  # Remove default sheet
        
        # Define styles for overview sheet
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, size=12, color="333333")
        
        # Track sheets created
        sheets_created = []
        total_sheets_processed = 0
        processing_log = []

        # Process each file
        for file_index, file_data in enumerate(file_data_list):
            try:
                # Read each Excel file
                file_content = file_data.get('content')
                original_name = file_data.get('originalName', f'file_{file_index}')
                file_type = file_data.get('fileType', 'unknown')
                
                if not file_content:
                    logger.warning(f"No content for file {original_name}")
                    processing_log.append(f"⚠️ Skipped {original_name}: No content")
                    continue
                    
                file_bytes = BytesIO(base64.b64decode(file_content))
                wb = openpyxl.load_workbook(file_bytes)
                
                logger.info(f"Processing {original_name} with {len(wb.sheetnames)} sheets")
                processing_log.append(f"✅ Processing {original_name}: {len(wb.sheetnames)} sheets")
                
                # Copy each sheet to the combined workbook
                for sheet_index, sheet_name in enumerate(wb.sheetnames):
                    source_sheet = wb[sheet_name]
                    
                    # Create unique sheet name
                    clean_original_name = re.sub(r'[^a-zA-Z0-9_]', '_', original_name.replace('.xlsx', ''))[:15]
                    clean_sheet_name = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name)[:15]
                    
                    # Create new sheet name with file prefix
                    new_sheet_name = f"{clean_original_name}_{clean_sheet_name}"
                    
                    # Ensure sheet name is within Excel limits (31 chars)
                    if len(new_sheet_name) > 31:
                        new_sheet_name = f"{clean_original_name[:12]}_{clean_sheet_name[:12]}_{file_index}"
                    
                    # Check for duplicates and make unique
                    counter = 1
                    original_new_sheet_name = new_sheet_name
                    while new_sheet_name in [sheet.title for sheet in combined_wb.worksheets]:
                        suffix = f"_{counter}"
                        max_base_length = 31 - len(suffix)
                        new_sheet_name = f"{original_new_sheet_name[:max_base_length]}{suffix}"
                        counter += 1
                    
                    new_sheet = combined_wb.create_sheet(title=new_sheet_name)
                    
                    # Copy all cell values and formatting
                    row_count = 0
                    for row in source_sheet.iter_rows():
                        has_data = False
                        for cell in row:
                            if cell.value is not None:
                                has_data = True
                                new_cell = new_sheet[cell.coordinate]
                                new_cell.value = cell.value
                                
                                # Copy formatting
                                try:
                                    if cell.font:
                                        new_cell.font = Font(
                                            name=cell.font.name or 'Calibri',
                                            size=cell.font.size or 11,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            color=cell.font.color
                                        )
                                    
                                    if cell.fill and cell.fill.fill_type:
                                        new_cell.fill = PatternFill(
                                            fill_type=cell.fill.fill_type,
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color
                                        )
                                    
                                    if cell.alignment:
                                        new_cell.alignment = Alignment(
                                            horizontal=cell.alignment.horizontal,
                                            vertical=cell.alignment.vertical,
                                            wrap_text=cell.alignment.wrap_text
                                        )
                                        
                                    if cell.border:
                                        new_cell.border = Border(
                                            left=cell.border.left,
                                            right=cell.border.right,
                                            top=cell.border.top,
                                            bottom=cell.border.bottom
                                        )
                                except Exception as formatting_error:
                                    # If formatting fails, just copy the value
                                    logger.warning(f"Formatting error for {cell.coordinate}: {str(formatting_error)}")
                        
                        if has_data:
                            row_count += 1
                    
                    # Copy column widths
                    try:
                        for col_letter, col_dimension in source_sheet.column_dimensions.items():
                            if col_dimension.width:
                                new_sheet.column_dimensions[col_letter].width = col_dimension.width
                            else:
                                # Set default width
                                new_sheet.column_dimensions[col_letter].width = 15
                    except Exception as e:
                        logger.warning(f"Column width copy error: {str(e)}")
                    
                    # Copy row heights
                    try:
                        for row_num, row_dimension in source_sheet.row_dimensions.items():
                            if row_dimension.height:
                                new_sheet.row_dimensions[row_num].height = row_dimension.height
                    except Exception as e:
                        logger.warning(f"Row height copy error: {str(e)}")
                    
                    # Set print settings
                    new_sheet.page_setup.orientation = new_sheet.ORIENTATION_LANDSCAPE
                    new_sheet.page_setup.paperSize = new_sheet.PAPERSIZE_A4
                    new_sheet.page_setup.fitToWidth = 1
                    new_sheet.page_setup.fitToHeight = 0
                    
                    sheets_created.append({
                        'name': new_sheet_name,
                        'source_file': original_name,
                        'source_sheet': sheet_name,
                        'rows_with_data': row_count
                    })
                    total_sheets_processed += 1
                    
                    logger.info(f"Added sheet: {new_sheet_name} from {original_name} ({row_count} rows)")
                    
            except Exception as e:
                error_msg = f"Error processing file {file_data.get('originalName', 'unknown')}: {str(e)}"
                logger.error(error_msg)
                processing_log.append(f"❌ {error_msg}")
                continue

        # If no sheets were created, create a message sheet
        if len(sheets_created) == 0:
            message_sheet = combined_wb.create_sheet("No_Data_Found")
            message_sheet['A1'] = f"No valid data found in {category_name} category files"
            message_sheet['A1'].font = Font(bold=True, size=14, color="FF0000")
            message_sheet['A3'] = f"Attempted to process: {len(file_data_list)} files"
            message_sheet['A4'] = f"Category: {category_name.replace('_', ' ').title()}"
            message_sheet['A5'] = f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Add processing log
            log_start_row = 7
            message_sheet[f'A{log_start_row}'] = "Processing Log:"
            message_sheet[f'A{log_start_row}'].font = Font(bold=True)
            for i, log_entry in enumerate(processing_log):
                message_sheet[f'A{log_start_row + 1 + i}'] = log_entry
            
            sheets_created.append({
                'name': 'No_Data_Found',
                'source_file': 'System Generated',
                'source_sheet': 'Error Message',
                'rows_with_data': len(processing_log) + 5
            })

        # Add category overview sheet at the beginning
        overview_sheet = combined_wb.create_sheet("Category_Overview", 0)
        
        # Overview sheet styling and content
        overview_sheet['A1'] = f"{category_name.upper().replace('_', ' ')} CATEGORY - COMBINED ANALYSIS"
        overview_sheet['A1'].font = title_font
        overview_sheet['A1'].fill = title_fill
        overview_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        overview_sheet.merge_cells('A1:F1')
        overview_sheet.row_dimensions[1].height = 30
        
        # Category summary section
        current_row = 3
        overview_sheet[f'A{current_row}'] = "📊 CATEGORY SUMMARY"
        overview_sheet[f'A{current_row}'].font = header_font
        current_row += 1
        
        summary_data = [
            ["Category Name:", category_name.replace('_', ' ').title()],
            ["Files Combined:", len(file_data_list)],
            ["Total Sheets Created:", len(sheets_created)],
            ["Sheets with Data:", len([s for s in sheets_created if s.get('rows_with_data', 0) > 0])],
            ["Fiscal Year:", fiscal_year or 'Not Specified'],
            ["Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S IST")],
            ["Combination Type:", combination_type.replace('_', ' ').title()]
        ]
        
        for label, value in summary_data:
            overview_sheet[f'A{current_row}'] = label
            overview_sheet[f'A{current_row}'].font = Font(bold=True)
            overview_sheet[f'B{current_row}'] = str(value)
            current_row += 1
        
        current_row += 2
        
        # Source files section
        overview_sheet[f'A{current_row}'] = "📁 SOURCE FILES"
        overview_sheet[f'A{current_row}'].font = header_font
        current_row += 1
        
        # Headers for source files table
        file_headers = ["#", "File Name", "Type", "Sheets Contributed"]
        for col, header in enumerate(file_headers):
            cell = overview_sheet.cell(row=current_row, column=col+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Source files data
        for i, file_data in enumerate(file_data_list):
            original_name = file_data.get('originalName', f'File_{i+1}')
            file_type = file_data.get('fileType', 'Unknown')
            sheets_from_file = len([s for s in sheets_created if s.get('source_file') == original_name])
            
            overview_sheet[f'A{current_row}'] = i + 1
            overview_sheet[f'B{current_row}'] = original_name
            overview_sheet[f'C{current_row}'] = file_type
            overview_sheet[f'D{current_row}'] = sheets_from_file
            current_row += 1
        
        current_row += 2
        
        # Created sheets section
        overview_sheet[f'A{current_row}'] = "📋 CREATED SHEETS DETAILS"
        overview_sheet[f'A{current_row}'].font = header_font
        current_row += 1
        
        # Headers for sheets table
        sheet_headers = ["#", "Sheet Name", "Source File", "Original Sheet", "Data Rows"]
        for col, header in enumerate(sheet_headers):
            cell = overview_sheet.cell(row=current_row, column=col+1, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1
        
        # Sheets data (excluding overview sheet)
        for i, sheet_info in enumerate(sheets_created):
            overview_sheet[f'A{current_row}'] = i + 1
            overview_sheet[f'B{current_row}'] = sheet_info.get('name', 'Unknown')
            overview_sheet[f'C{current_row}'] = sheet_info.get('source_file', 'Unknown')
            overview_sheet[f'D{current_row}'] = sheet_info.get('source_sheet', 'Unknown')
            overview_sheet[f'E{current_row}'] = sheet_info.get('rows_with_data', 0)
            current_row += 1
        
        current_row += 2
        
        # Processing log section
        if processing_log:
            overview_sheet[f'A{current_row}'] = "📜 PROCESSING LOG"
            overview_sheet[f'A{current_row}'].font = header_font
            current_row += 1
            
            for log_entry in processing_log:
                overview_sheet[f'A{current_row}'] = log_entry
                overview_sheet[f'A{current_row}'].font = Font(size=10)
                current_row += 1
        
        # Auto-adjust column widths for overview sheet
        for column in overview_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 3, 60)
            overview_sheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        # Add borders to overview sheet tables
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to summary section
        for row in range(4, 11):
            for col in range(1, 3):
                overview_sheet.cell(row=row, column=col).border = thin_border
        
        # Apply borders to files table
        files_table_start = 14
        files_table_end = files_table_start + len(file_data_list)
        for row in range(files_table_start, files_table_end + 1):
            for col in range(1, 5):
                overview_sheet.cell(row=row, column=col).border = thin_border
        
        # Apply borders to sheets table
        sheets_table_start = files_table_end + 4
        sheets_table_end = sheets_table_start + len(sheets_created)
        for row in range(sheets_table_start, sheets_table_end + 1):
            for col in range(1, 6):
                overview_sheet.cell(row=row, column=col).border = thin_border
        
        # Set print settings for overview sheet
        overview_sheet.page_setup.orientation = overview_sheet.ORIENTATION_LANDSCAPE
        overview_sheet.page_setup.paperSize = overview_sheet.PAPERSIZE_A4
        overview_sheet.page_setup.fitToWidth = 1
        overview_sheet.page_setup.fitToHeight = 0
        
        # Freeze panes at row 3 for overview
        overview_sheet.freeze_panes = 'A3'

        # Update sheets_created to include overview
        final_sheets = ["Category_Overview"] + [s['name'] for s in sheets_created]

        # Save combined workbook
        output = BytesIO()
        combined_wb.save(output)
        output.seek(0)
        combined_data = output.getvalue()

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{category_name.replace('_', '-')}_combined_{fiscal_year}_{timestamp}.xlsx"

        # Generate comprehensive metadata
        file_metadata = {
            'category': category_name,
            'category_display_name': category_name.replace('_', ' ').title(),
            'files_combined': len(file_data_list),
            'sheets_created': final_sheets,
            'total_sheets': len(final_sheets),
            'data_sheets': len(sheets_created),
            'fiscal_year': fiscal_year,
            'combination_type': combination_type,
            'generated_on': datetime.now().isoformat(),
            'file_size': len(combined_data),
            'source_files': [f.get('originalName') for f in file_data_list],
            'processing_log': processing_log,
            'sheets_details': sheets_created,
            'total_data_rows': sum(s.get('rows_with_data', 0) for s in sheets_created),
            'has_overview_sheet': True,
            'overview_sheet_name': 'Category_Overview'
        }

        logger.info(f"Successfully combined {len(file_data_list)} files from {category_name} category into {len(final_sheets)} total sheets")

        return jsonify({
            'success': True,
            'file_data': base64.b64encode(combined_data).decode('utf-8'),
            'file_name': file_name,
            'metadata': file_metadata,
            'message': f'Successfully combined {len(file_data_list)} files with {len(sheets_created)} data sheets plus overview'
        })

    except Exception as e:
        logger.error(f"Error combining category Excel files: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Failed to combine category Excel files: {str(e)}'
        }),500

@combined_bp.route('/combine-excel-files', methods=['POST'])
def combine_excel_files():
    """Combine multiple Excel files into one master file with specific sheet order and names"""
    try:
        data = request.get_json()
        file_data_list = data.get('files', [])
        excel_formatting = data.get('excel_formatting', {})
        metadata = data.get('metadata', {})
        
        # NEW: Add title option parameter
        # Options: 'none' (no titles), 'full' (full names), 'short' (current behavior)
        title_option = data.get('title_option', 'short')  # Default to current behavior
        
        if not file_data_list:
            return jsonify({
                'success': False,
                'error': 'No files provided for combining'
            }), 400

        logger.info(f"Combining {len(file_data_list)} files with title option: {title_option}")

        # Create a new workbook with specific sheet order
        combined_wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in combined_wb.sheetnames:
            combined_wb.remove(combined_wb['Sheet'])
        
        # Define the sheet order and names
        sheet_order = [
            "Sales Analysis Month wise",
            "Region wise analysis",
            "Product wise analysis",
            "TS-PW",
            "ERO-PW"
        ]
        
        # Create empty sheets in the specified order
        for sheet_name in sheet_order:
            combined_wb.create_sheet(sheet_name)
        
        # Define styles for source file titles and table headers
        source_title_font = Font(bold=True, size=12, color="FFFFFF")
        source_title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        source_title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Table header styles
        table_header_font = Font(bold=True, size=11, color="FFFFFF")
        table_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        table_header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data cell alignment
        data_alignment_center = Alignment(horizontal="center", vertical="center")
        data_alignment_left = Alignment(horizontal="left", vertical="center")
        data_alignment_right = Alignment(horizontal="right", vertical="center")
        
        # Track sheets created and processing log
        sheets_created = []
        processing_log = []
        category_breakdown = {}

        # Process each file and distribute data to appropriate sheets
        for file_index, file_data in enumerate(file_data_list):
            try:
                file_content = file_data.get('content')
                file_name = file_data.get('name', f'file_{file_index}')
                category = file_data.get('category', 'unknown')
                
                if not file_content:
                    logger.warning(f"No content for file {file_name}")
                    processing_log.append(f"⚠️ Skipped {file_name}: No content")
                    continue
                
                # Decode base64 content
                file_bytes = BytesIO(base64.b64decode(file_content))
                wb = openpyxl.load_workbook(file_bytes, data_only=False)
                
                logger.info(f"Processing {file_name} with {len(wb.sheetnames)} sheets")
                processing_log.append(f"✅ Processing {file_name}: {len(wb.sheetnames)} sheets")
                
                category_breakdown[category] = category_breakdown.get(category, 0) + len(wb.sheetnames)
                
                # Determine target sheet based on file content/category
                target_sheet_name = None
                if 'sales' in file_name.lower() or 'month' in file_name.lower():
                    target_sheet_name = sheet_order[0]  # Sales Analysis Monthwise
                elif 'region' in file_name.lower():
                    target_sheet_name = sheet_order[1]  # Region Wise Analysis
                elif 'product' in file_name.lower():
                    target_sheet_name = sheet_order[2]  # Product Wise Analysis
                elif 'ts' in file_name.lower() or 'ts-pw' in file_name.lower():
                    target_sheet_name = sheet_order[3]  # TS-PW
                elif 'ero' in file_name.lower() or 'ero-pw' in file_name.lower():
                    target_sheet_name = sheet_order[4]  # ERO-PW
                
                if not target_sheet_name:
                    logger.warning(f"No matching sheet for {file_name}")
                    processing_log.append(f"⚠️ No target sheet for {file_name}")
                    continue
                
                target_sheet = combined_wb[target_sheet_name]
                
                # Copy data from all sheets in the file to the target sheet
                for source_sheet_name in wb.sheetnames:
                    source_sheet = wb[source_sheet_name]
                    
                    # Find the next empty row in target sheet
                    max_row = target_sheet.max_row
                    if max_row == 1 and target_sheet.cell(row=1, column=1).value is None:
                        start_row = 1
                    else:
                        # Adjust spacing based on title option
                        if title_option == 'none':
                            start_row = max_row + 2  # Smaller gap when no title
                        else:
                            start_row = max_row + 3  # Leave space for title
                    
                    # Add source file title based on option
                    if title_option != 'none' and start_row > 1:
                        # Generate title based on option
                        if title_option == 'full':
                            # Full title with complete file name and sheet name
                            clean_file_name = file_name.replace('.xlsx', '').replace('.xls', '')
                            title_text = f"📊 {clean_file_name} - {source_sheet_name}"
                        elif title_option == 'short':
                            # Current behavior - shortened title
                            clean_file_name = file_name.replace('.xlsx', '').replace('.xls', '')
                            if len(clean_file_name) > 30:
                                clean_file_name = clean_file_name[:30] + "..."
                            title_text = f"📊 {clean_file_name} - {source_sheet_name}"
                        
                        title_row = start_row - 2
                        title_cell = target_sheet.cell(row=title_row, column=1, value=title_text)
                        title_cell.font = source_title_font
                        title_cell.fill = source_title_fill
                        title_cell.alignment = source_title_alignment
                        
                        # FIXED: Calculate merge range for EACH table title separately
                        # Get the actual data width from the source sheet
                        source_max_col = source_sheet.max_column or 5
                        
                        # Calculate merge range based on title text length and ensure full visibility
                        title_char_count = len(title_text)
                        
                        # More aggressive calculation to ensure no cutoff
                        if title_option == 'full':
                            # For full titles, use very generous merge range
                            title_length_cols = max(title_char_count, 40)  # Direct character count, minimum 40
                            merge_end_col = max(source_max_col, int(title_length_cols), 45)  # Minimum 45 columns
                        else:  # short
                            # For short titles, still be very generous
                            title_length_cols = max(title_char_count, 30)  # Direct character count, minimum 30
                            merge_end_col = max(source_max_col, int(title_length_cols), 35)  # Minimum 35 columns
                        
                        # EXTRA special handling for shorter sheet names (TS-PW, ERO-PW)
                        if target_sheet_name in ['TS-PW', 'ERO-PW']:
                            # These sheets need MUCH more space - force wide merge
                            merge_end_col = max(merge_end_col, 50)  # Force minimum 50 columns for these sheets
                            logger.info(f"Applied special handling for sheet '{target_sheet_name}' - forced merge_end_col to {merge_end_col}")
                        
                        # Allow very wide merge ranges to prevent any cutoff
                        merge_end_col = min(merge_end_col, 100)  # Increased cap to 100 columns
                        
                        logger.info(f"Sheet: '{target_sheet_name}' | Title: '{title_text}' (chars: {title_char_count}) -> merge_end_col: {merge_end_col}")
                        
                        # IMPORTANT: Check if merge range conflicts with existing merged cells
                        try:
                            # Unmerge any existing cells in this range first
                            ranges_to_unmerge = []
                            for merged_range in target_sheet.merged_cells.ranges:
                                if (merged_range.min_row == title_row and 
                                    merged_range.min_col <= merge_end_col and 
                                    merged_range.max_col >= 1):
                                    ranges_to_unmerge.append(merged_range)
                            
                            for range_to_unmerge in ranges_to_unmerge:
                                target_sheet.unmerge_cells(str(range_to_unmerge))
                            
                            # Now merge the title cells
                            merge_range = f"A{title_row}:{openpyxl.utils.get_column_letter(merge_end_col)}{title_row}"
                            target_sheet.merge_cells(merge_range)
                            
                            logger.info(f"Merged title range: {merge_range} for '{title_text}'")
                            
                        except Exception as merge_error:
                            logger.warning(f"Could not merge cells for title: {merge_error}")
                            # If merge fails, at least ensure the title is visible
                            title_cell.alignment = source_title_alignment
                        
                        # Set row height for title
                        if title_option == 'full':
                            target_sheet.row_dimensions[title_row].height = 40  # Taller for full titles
                        else:
                            target_sheet.row_dimensions[title_row].height = 30
                        
                        # Apply title formatting to the entire merged range - ENSURE ALL CELLS GET FORMATTING
                        for col in range(1, merge_end_col + 1):
                            try:
                                cell = target_sheet.cell(row=title_row, column=col)
                                cell.font = source_title_font
                                cell.fill = source_title_fill
                                cell.alignment = source_title_alignment
                                # Force the value in the first cell only
                                if col == 1:
                                    cell.value = title_text
                            except Exception as format_error:
                                logger.warning(f"Could not format cell {col} in title row: {format_error}")
                        
                        # ADDITIONAL: Force wider columns for the title area
                        if target_sheet_name in ['TS-PW', 'ERO-PW']:
                            # Force first 10 columns to be wider for these problematic sheets
                            for col_num in range(1, min(11, merge_end_col + 1)):
                                col_letter = openpyxl.utils.get_column_letter(col_num)
                                current_width = target_sheet.column_dimensions[col_letter].width or 8.43
                                target_sheet.column_dimensions[col_letter].width = max(current_width, 15)
                                logger.info(f"Set column {col_letter} width to {target_sheet.column_dimensions[col_letter].width} for sheet {target_sheet_name}")
                    
                    # Copy data with formatting and apply center alignment to headers
                    for row_idx, row in enumerate(source_sheet.iter_rows(), 1):
                        is_header_row = False
                        
                        # Check if this is likely a header row
                        if row_idx <= 3:  # First 3 rows might be headers
                            row_values = [str(cell.value or '').upper() for cell in row if cell.value]
                            header_keywords = ['SALES', 'ORGANIZATION', 'BUDGET', 'ACTUAL', 'MONTH', 'VALUE', 'MT', 'NAME', 'TOTAL']
                            is_header_row = any(keyword in ' '.join(row_values) for keyword in header_keywords)
                        
                        for cell in row:
                            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                                continue
                                
                            if cell.value is not None:
                                new_cell = target_sheet.cell(
                                    row=cell.row + start_row - 1,
                                    column=cell.column,
                                    value=cell.value
                                )
                                
                                # Copy formatting
                                try:
                                    if cell.font:
                                        new_cell.font = Font(
                                            name=cell.font.name or 'Calibri',
                                            size=cell.font.size or 11,
                                            bold=cell.font.bold,
                                            italic=cell.font.italic,
                                            color=cell.font.color
                                        )
                                    
                                    if cell.fill and cell.fill.fill_type:
                                        new_cell.fill = PatternFill(
                                            fill_type=cell.fill.fill_type,
                                            start_color=cell.fill.start_color,
                                            end_color=cell.fill.end_color
                                        )
                                    
                                    # Apply center alignment for headers, keep original for data
                                    if is_header_row:
                                        new_cell.alignment = table_header_alignment
                                        # Make headers bold and with header styling
                                        new_cell.font = Font(
                                            name=new_cell.font.name or 'Calibri',
                                            size=new_cell.font.size or 11,
                                            bold=True,
                                            color=new_cell.font.color or "000000"
                                        )
                                    else:
                                        # Apply appropriate alignment for data cells
                                        if cell.alignment:
                                            new_cell.alignment = Alignment(
                                                horizontal=cell.alignment.horizontal or "center",
                                                vertical="center",
                                                wrap_text=cell.alignment.wrap_text
                                            )
                                        else:
                                            # Default alignment based on data type
                                            if isinstance(cell.value, (int, float)):
                                                new_cell.alignment = data_alignment_right
                                            elif isinstance(cell.value, str) and cell.column == 1:
                                                new_cell.alignment = data_alignment_left
                                            else:
                                                new_cell.alignment = data_alignment_center
                                    
                                    if cell.border:
                                        new_cell.border = Border(
                                            left=cell.border.left,
                                            right=cell.border.right,
                                            top=cell.border.top,
                                            bottom=cell.border.bottom
                                        )
                                    
                                    if cell.number_format:
                                        new_cell.number_format = cell.number_format
                                except Exception as e:
                                    logger.warning(f"Formatting copy error: {str(e)}")
                    
                    sheets_created.append({
                        'target_sheet': target_sheet_name,
                        'source_file': file_name,
                        'source_sheet': source_sheet_name,
                        'rows_added': source_sheet.max_row,
                        'title_option': title_option,
                        'merge_range_used': f"A{title_row}:{openpyxl.utils.get_column_letter(merge_end_col)}{title_row}" if title_option != 'none' else None
                    })
                    
                    logger.info(f"Added data from {file_name} to {target_sheet_name}")
                    
            except Exception as e:
                error_msg = f"Error processing file {file_name}: {str(e)}"
                logger.error(error_msg)
                processing_log.append(f"❌ {error_msg}")
                continue

        # Auto-adjust column widths for all sheets BEFORE applying final formatting
        for sheet in combined_wb.worksheets:
            # Calculate column widths based on content
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set width with appropriate limits based on title option
                if column_letter == 'A':
                    # Column A needs to accommodate titles
                    if title_option == 'full':
                        min_width = max(max_length + 5, 60)  # Wider for full titles
                    elif title_option == 'short':
                        min_width = max(max_length + 3, 40)  # Current width
                    else:  # 'none'
                        min_width = max(max_length + 2, 25)  # Narrower when no titles
                    
                    sheet.column_dimensions[column_letter].width = min(min_width, 100)
                else:
                    # Other columns
                    if max_length > 0:
                        adjusted_width = min(max_length + 3, 50)
                        sheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
                    else:
                        sheet.column_dimensions[column_letter].width = 15

        # Apply final formatting to all sheets - only if titles are enabled
        if title_option != 'none':
            for sheet in combined_wb.worksheets:
                # Find and format all title rows and header rows
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            cell_value = str(cell.value).upper()
                            
                            # Check if this is a title row (contains source file info)
                            if '📊' in str(cell.value) or 'DATA FROM:' in cell_value:
                                cell.alignment = source_title_alignment
                                cell.font = source_title_font
                                if not cell.fill or cell.fill.fill_type != 'solid':
                                    cell.fill = source_title_fill
                            
                            # Check if this is a table header row
                            elif any(keyword in cell_value for keyword in ['SALES', 'ORGANIZATION', 'BUDGET', 'ACTUAL', 'MT', 'VALUE']):
                                if cell.row <= 15:  # Increased range for headers (first 15 rows)
                                    cell.alignment = table_header_alignment
                                    # Make header bold if not already
                                    if not cell.font.bold:
                                        cell.font = Font(
                                            name=cell.font.name or 'Calibri',
                                            size=cell.font.size or 11,
                                            bold=True,
                                            color=cell.font.color
                                        )

        # Set print settings for all sheets
        for sheet in combined_wb.worksheets:
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0

        # Save combined workbook
        output = BytesIO()
        combined_wb.save(output)
        output.seek(0)
        combined_data = output.getvalue()

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Auditor_Format_{timestamp}.xlsx"

        # Generate metadata
        file_metadata = {
            'files_combined': len(file_data_list),
            'sheets_order': sheet_order,
            'total_sheets': len(sheet_order),
            'category_breakdown': category_breakdown,
            'generated_on': datetime.now().isoformat(),
            'file_size': len(combined_data),
            'source_files': [f.get('name') for f in file_data_list],
            'processing_log': processing_log,
            'sheets_details': sheets_created,
            'title_option_used': title_option  # Track which option was used
        }

        logger.info(f"Successfully created combined file with {len(sheet_order)} sheets using title option: {title_option}")

        return jsonify({
            'success': True,
            'file_data': base64.b64encode(combined_data).decode('utf-8'),
            'file_name': file_name,
            'metadata': file_metadata,
            'message': f'Successfully combined files into {len(sheet_order)} sheets with title option: {title_option}'
        })

    except Exception as e:
        logger.error(f"Error combining Excel files: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': f'Failed to combine Excel files: {str(e)}'
        }), 500

@combined_bp.route('/get-integration-status', methods=['GET'])
def get_integration_status():
    """Get integration status for all analysis types"""
    try:
        # Check session for integration data
        integration_status = {
            'sales_analysis': {
                'registered': session.get('sales_analysis_integrated', False),
                'files_count': session.get('sales_analysis_files_count', 0),
                'last_update': session.get('sales_analysis_last_update'),
                'records': {
                    'mt': session.get('sales_analysis_mt_records', 0),
                    'value': session.get('sales_analysis_value_records', 0)
                }
            },
            'region_analysis': {
                'registered': session.get('region_analysis_integrated', False),
                'files_count': session.get('region_analysis_files_count', 0),
                'last_update': session.get('region_analysis_last_update'),
                'records': {
                    'mt': session.get('region_analysis_mt_records', 0),
                    'value': session.get('region_analysis_value_records', 0)
                }
            },
            'ero_pw_analysis': {
                'registered': session.get('ero_pw_integrated', False),
                'files_count': session.get('ero_pw_files_count', 0),
                'last_update': session.get('ero_pw_last_update'),
                'records': {
                    'mt': session.get('ero_pw_mt_records', 0),
                    'value': session.get('ero_pw_value_records', 0)
                }
            },
            'ts_pw_analysis': {
                'registered': session.get('ts_pw_integrated', False),
                'files_count': session.get('ts_pw_files_count', 0),
                'last_update': session.get('ts_pw_last_update'),
                'records': {
                    'mt': session.get('ts_pw_mt_records', 0),
                    'value': session.get('ts_pw_value_records', 0)
                }
            },
            'product_analysis': {
                'registered': session.get('product_analysis_integrated', False),
                'files_count': session.get('product_analysis_files_count', 0),
                'last_update': session.get('product_analysis_last_update'),
                'records': {
                    'mt': session.get('product_analysis_mt_records', 0),
                    'value': session.get('product_analysis_value_records', 0)
                }
            }
        }
        
        # Check file storage callbacks
        active_callbacks = len([c for c in FILE_STORAGE_CALLBACKS.values() if c['status'] == 'active'])
        
        return jsonify({
            'success': True,
            'integration_status': integration_status,
            'active_storage_callbacks': active_callbacks,
            'total_registered_analyses': sum(1 for status in integration_status.values() if status['registered'])
        })
        
    except Exception as e:
        logger.error(f"Error getting integration status: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to get integration status: {str(e)}'
        }), 500

@combined_bp.route('/clear-integration-cache', methods=['POST'])
def clear_integration_cache():
    """Clear all integration cache and callbacks"""
    try:
        # Clear session data for all analysis types
        analysis_types = ['sales_analysis', 'region_analysis', 'ero_pw', 'ts_pw', 'product_analysis']
        cleared_items = []
        
        for analysis_type in analysis_types:
            keys_to_clear = [
                f'{analysis_type}_integrated',
                f'{analysis_type}_files_count',
                f'{analysis_type}_last_update',
                f'{analysis_type}_mt_records',
                f'{analysis_type}_value_records'
            ]
            
            for key in keys_to_clear:
                if key in session:
                    session.pop(key, None)
                    cleared_items.append(key)
        
        # Clear all file storage callbacks
        callback_count = len(FILE_STORAGE_CALLBACKS)
        FILE_STORAGE_CALLBACKS.clear()
        
        logger.info(f"Integration cache cleared: {len(cleared_items)} session items, {callback_count} callbacks")
        
        return jsonify({
            'success': True,
            'message': 'Integration cache and callbacks cleared successfully',
            'cleared_session_items': len(cleared_items),
            'cleared_callbacks': callback_count
        })
        
    except Exception as e:
        logger.error(f"Error clearing integration cache: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to clear integration cache: {str(e)}'
        }), 500


# Export the blueprint
__all__ = ['combined_bp']