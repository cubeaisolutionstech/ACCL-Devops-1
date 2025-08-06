# Create this file: utilities/excel_formatting.py

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, GradientFill
import logging

logger = logging.getLogger(__name__)

def copy_cell_comprehensive_formatting(source_cell, target_cell, excel_formatting=None):
    """
    Copy comprehensive formatting from source cell to target cell
    """
    try:
        # Copy basic value
        target_cell.value = source_cell.value
        
        # Font formatting with all properties
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name or 'Calibri',
                size=source_cell.font.size or 11,
                bold=source_cell.font.bold or False,
                italic=source_cell.font.italic or False,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline or 'none',
                strike=source_cell.font.strike or False,
                color=source_cell.font.color,
                charset=source_cell.font.charset,
                family=source_cell.font.family,
                scheme=source_cell.font.scheme
            )
        
        # Fill/Background formatting with gradient support
        if source_cell.fill and source_cell.fill.fill_type:
            if source_cell.fill.fill_type == 'solid':
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            elif source_cell.fill.fill_type in ['linear', 'path']:
                # Handle gradient fills
                try:
                    target_cell.fill = GradientFill(
                        type=source_cell.fill.type,
                        degree=getattr(source_cell.fill, 'degree', 0),
                        left=getattr(source_cell.fill, 'left', 0),
                        right=getattr(source_cell.fill, 'right', 0),
                        top=getattr(source_cell.fill, 'top', 0),
                        bottom=getattr(source_cell.fill, 'bottom', 0),
                        stop=getattr(source_cell.fill, 'stop', [])
                    )
                except:
                    # Fallback to pattern fill if gradient fails
                    target_cell.fill = PatternFill(
                        fill_type='solid',
                        start_color=source_cell.fill.start_color,
                        end_color=source_cell.fill.end_color
                    )
        
        # Enhanced alignment formatting
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation or 0,
                wrap_text=source_cell.alignment.wrap_text or False,
                shrink_to_fit=source_cell.alignment.shrink_to_fit or False,
                indent=source_cell.alignment.indent or 0,
                relativeIndent=getattr(source_cell.alignment, 'relativeIndent', 0),
                justifyLastLine=getattr(source_cell.alignment, 'justifyLastLine', False),
                readingOrder=getattr(source_cell.alignment, 'readingOrder', 0)
            )
        
        # Comprehensive border formatting
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom,
                diagonal=source_cell.border.diagonal,
                diagonal_direction=source_cell.border.diagonal_direction,
                outline=getattr(source_cell.border, 'outline', True),
                vertical=getattr(source_cell.border, 'vertical', None),
                horizontal=getattr(source_cell.border, 'horizontal', None)
            )
        
        # Enhanced number formatting with custom format preservation
        if source_cell.number_format and source_cell.number_format != 'General':
            target_cell.number_format = source_cell.number_format
        elif isinstance(source_cell.value, (int, float)) and not pd.isna(source_cell.value):
            # Apply formatting based on excel_formatting configuration
            if excel_formatting and excel_formatting.get('number_format', {}).get('force_format', False):
                format_pattern = excel_formatting.get('number_format', {}).get('format_pattern', '0.00')
                target_cell.number_format = format_pattern
            elif source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            else:
                # Default number formatting
                target_cell.number_format = '0.00'
        
        # Copy hyperlinks
        if hasattr(source_cell, 'hyperlink') and source_cell.hyperlink:
            target_cell.hyperlink = source_cell.hyperlink
        
        # Copy comments with formatting
        if hasattr(source_cell, 'comment') and source_cell.comment:
            target_cell.comment = source_cell.comment
        
        return True
        
    except Exception as e:
        logger.warning(f"Error copying comprehensive formatting for cell {source_cell.coordinate}: {str(e)}")
        # Fallback to basic value copy
        target_cell.value = source_cell.value
        return False

def copy_sheet_structural_formatting(source_sheet, target_sheet):
    """
    Copy structural formatting elements from source sheet to target sheet
    """
    try:
        # Copy column dimensions with enhanced precision
        for col_letter, col_dimension in source_sheet.column_dimensions.items():
            target_col_dim = target_sheet.column_dimensions[col_letter]
            
            # Copy width
            if col_dimension.width:
                target_col_dim.width = col_dimension.width
            
            # Copy visibility
            if hasattr(col_dimension, 'hidden') and col_dimension.hidden:
                target_col_dim.hidden = col_dimension.hidden
        
        # Copy row dimensions with enhanced properties
        for row_num, row_dimension in source_sheet.row_dimensions.items():
            target_row_dim = target_sheet.row_dimensions[row_num]
            
            # Copy height
            if row_dimension.height:
                target_row_dim.height = row_dimension.height
            
            # Copy visibility
            if hasattr(row_dimension, 'hidden') and row_dimension.hidden:
                target_row_dim.hidden = row_dimension.hidden
        
        # Copy merged cells ranges
        for merged_range in source_sheet.merged_cells.ranges:
            try:
                target_sheet.merge_cells(str(merged_range))
            except Exception as merge_error:
                logger.warning(f"Could not merge cells {merged_range}: {str(merge_error)}")
        
        # Copy freeze panes
        if source_sheet.freeze_panes:
            target_sheet.freeze_panes = source_sheet.freeze_panes
        
        # Copy print settings
        try:
            target_sheet.page_setup.orientation = source_sheet.page_setup.orientation
            target_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
            target_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
            target_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
            
            # Copy margins
            if source_sheet.page_margins:
                target_sheet.page_margins.left = source_sheet.page_margins.left
                target_sheet.page_margins.right = source_sheet.page_margins.right
                target_sheet.page_margins.top = source_sheet.page_margins.top
                target_sheet.page_margins.bottom = source_sheet.page_margins.bottom
                
        except Exception as print_error:
            logger.warning(f"Could not copy print settings: {str(print_error)}")
            
        return True
        
    except Exception as e:
        logger.error(f"Error copying structural formatting: {str(e)}")
        return False

def apply_intelligent_number_formatting(worksheet, excel_formatting=None):
    """
    Apply intelligent number formatting to a worksheet based on content
    """
    if not excel_formatting:
        return
    
    number_format_config = excel_formatting.get('number_format', {})
    if not number_format_config.get('apply_to_all_numbers', False):
        return
    
    format_pattern = number_format_config.get('format_pattern', '0.00')
    force_format = number_format_config.get('force_format', False)
    
    try:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    # Check if cell already has a specific number format
                    if force_format or cell.number_format == 'General':
                        cell.number_format = format_pattern
                        
        logger.info(f"Applied intelligent number formatting '{format_pattern}' to worksheet {worksheet.title}")
        
    except Exception as e:
        logger.warning(f"Could not apply intelligent number formatting: {str(e)}")

def detect_and_preserve_callback_formatting(file_metadata):
    """
    Detect if a file was generated via callback and should preserve special formatting
    """
    if not file_metadata:
        return False
    
    # Check for callback indicators
    callback_indicators = [
        file_metadata.get('wasCallbackStored', False),
        file_metadata.get('autoStored', False),
        file_metadata.get('originalFormatting', False),
        'callback' in str(file_metadata.get('source', '')).lower(),
        file_metadata.get('formattingApplied', False)
    ]
    
    return any(callback_indicators)

def enhance_overview_sheet_formatting(overview_sheet):
    """
    Apply enhanced formatting to the overview sheet
    """
    try:
        # Define enhanced color scheme
        colors = {
            'primary': "1F4E79",      # Dark blue
            'secondary': "4472C4",     # Medium blue  
            'success': "28A745",       # Green
            'info': "17A2B8",          # Teal
            'warning': "FFC107",       # Yellow
            'light': "F8F9FA",         # Light gray
            'white': "FFFFFF"          # White
        }
        
        # Add borders to all data cells
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for row in overview_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
        
        logger.info("Enhanced overview sheet formatting applied successfully")
        
    except Exception as e:
        logger.warning(f"Could not enhance overview sheet formatting: {str(e)}")